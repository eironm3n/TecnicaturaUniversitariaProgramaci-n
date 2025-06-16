import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

# Base de datos
conn = sqlite3.connect('estacionamiento.db')
cursor = conn.cursor()
cursor.execute('''CREATE TABLE IF NOT EXISTS vehiculos (
    patente TEXT,
    sector TEXT,
    ingreso TEXT,
    egreso TEXT,
    tiempo_estimado TEXT,
    tiempo_total TEXT,
    precio_hora REAL,
    precio_total REAL,
    pago_confirmado INTEGER DEFAULT 0
)''')
conn.commit()

# Excel
excel_file = "registro_estacionamiento.xlsx"
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["Patente", "Sector", "Ingreso", "Egreso", "Tiempo estimado", "Tiempo total", "Precio por Hora", "Precio Total"])
    wb.save(excel_file)

def actualizar_excel():
    wb = load_workbook(excel_file)
    ws = wb.active
    ws.delete_rows(2, ws.max_row)
    cursor.execute("SELECT * FROM vehiculos")
    for row in cursor.fetchall():
        ws.append(row)
    wb.save(excel_file)

def registrar():
    patente = entry_patente.get()
    sector = combo_sector.get()
    tiempo_est = entry_tiempo.get()
    ingreso = entry_ingreso.get()
    egreso = entry_egreso.get()

    if not patente or not sector:
        messagebox.showerror("Error", "Patente y Sector son obligatorios.")
        return

    cursor.execute("INSERT INTO vehiculos (patente, sector, ingreso, egreso, tiempo_estimado, tiempo_total, precio_hora, precio_total) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                   (patente, sector, ingreso, egreso, tiempo_est, '', 0.0, 0.0))
    conn.commit()
    actualizar_excel()
    mostrar_datos()
    entry_patente.delete(0, tk.END)
    entry_tiempo.delete(0, tk.END)
    entry_ingreso.delete(0, tk.END)
    entry_egreso.delete(0, tk.END)

def mostrar_datos():
    for item in tree.get_children():
        tree.delete(item)
    cursor.execute("SELECT * FROM vehiculos")
    for row in cursor.fetchall():
        color = "turquoise" if row[3] else "white"
        tree.insert("", tk.END, values=row, tags=(color,))
    tree.tag_configure("turquoise", background="turquoise")
    tree.tag_configure("white", background="white")

def actualizar_reloj():
    hora_actual = datetime.now().strftime("%H:%M:%S")
    reloj_label.config(text=hora_actual)
    reloj_label.after(1000, actualizar_reloj)

# Ventana principal
root = tk.Tk()
root.title("Sistema de Estacionamiento")
root.geometry("1200x650")

# Marco contenedor principal (rectángulo)
frame_principal = tk.Frame(root, bd=2, relief="groove")
frame_principal.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

# Dividir frame_principal verticalmente en dos: cabecera (30%), cuerpo (70%)
frame_principal.rowconfigure(0, weight=3)  # 30%
frame_principal.rowconfigure(1, weight=7)  # 70%
frame_principal.columnconfigure(0, weight=1)

# --- CABECERA ---
frame_cabecera = tk.Frame(frame_principal)
frame_cabecera.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

# Dividir cabecera horizontalmente: zona izquierda 70%, zona derecha 30%
frame_cabecera.columnconfigure(0, weight=7)
frame_cabecera.columnconfigure(1, weight=3)
frame_cabecera.rowconfigure(0, weight=1)

# Zona izquierda inputs (dos bloques organizados horizontalmente)
frame_inputs = tk.Frame(frame_cabecera)
frame_inputs.grid(row=0, column=0, sticky="nsew")

# Dividir inputs horizontalmente en dos bloques
frame_inputs.columnconfigure(0, weight=1)
frame_inputs.columnconfigure(1, weight=1)

# Bloque 1 (Patente, Sector, Tiempo estimado)
frame_bloque1 = tk.Frame(frame_inputs)
frame_bloque1.grid(row=0, column=0, sticky="w", padx=(0, 20))

tk.Label(frame_bloque1, text="Patente", font=("Arial", 10)).grid(row=0, column=0, padx=5, pady=2)
tk.Label(frame_bloque1, text="Sector de estacionamiento", font=("Arial", 10)).grid(row=0, column=1, padx=5, pady=2)
tk.Label(frame_bloque1, text="Tiempo estimado", font=("Arial", 10)).grid(row=0, column=2, padx=5, pady=2)

entry_patente = tk.Entry(frame_bloque1, width=12, font=("Arial", 12))
entry_patente.grid(row=1, column=0, padx=5)
entry_patente.insert(0, "")

combo_sector = ttk.Combobox(frame_bloque1, values=["Sector 1", "Sector 2", "Sector 3", "Sector 4"], width=14, font=("Arial", 12))
combo_sector.grid(row=1, column=1, padx=5)
combo_sector.set("")

entry_tiempo = tk.Entry(frame_bloque1, width=12, font=("Arial", 12))
entry_tiempo.grid(row=1, column=2, padx=5)
entry_tiempo.insert(0, "")

# Bloque 2 (Ingreso, Egreso)
frame_bloque2 = tk.Frame(frame_inputs)
frame_bloque2.grid(row=0, column=1, sticky="w")

tk.Label(frame_bloque2, text="Ingreso", font=("Arial", 10)).grid(row=0, column=0, padx=5, pady=2)
tk.Label(frame_bloque2, text="Egreso", font=("Arial", 10)).grid(row=0, column=1, padx=5, pady=2)

entry_ingreso = tk.Entry(frame_bloque2, width=15, font=("Arial", 12))
entry_ingreso.grid(row=1, column=0, padx=5)
entry_ingreso.insert(0, "")

entry_egreso = tk.Entry(frame_bloque2, width=15, font=("Arial", 12))
entry_egreso.grid(row=1, column=1, padx=5)
entry_egreso.insert(0, "")

# Botón Registrar debajo de bloque 2, alineado a la derecha dentro de inputs
btn_registrar = tk.Button(frame_inputs, text="Registrar", command=registrar)
btn_registrar.grid(row=1, column=1, sticky="e", padx=5, pady=5)

# Zona derecha: Reloj digital grande moderno
reloj_label = tk.Label(frame_cabecera, text="", font=("Segoe UI", 30, "bold"), fg="white", bg="black", padx=20, pady=10)
reloj_label.grid(row=0, column=1, sticky="ne", padx=10, pady=10)
actualizar_reloj()

# --- CUERPO ---
frame_cuerpo = tk.Frame(frame_principal)
frame_cuerpo.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
frame_cuerpo.rowconfigure(0, weight=1)
frame_cuerpo.columnconfigure(0, weight=1)

# Tabla de visualización
columns = ("patente", "sector", "ingreso", "egreso", "tiempo_estimado", "tiempo_total", "precio_hora", "precio_total")
tree = ttk.Treeview(frame_cuerpo, columns=columns, show="headings")
tree.grid(row=0, column=0, sticky="nsew")

# Scroll vertical para la tabla
scrollbar = ttk.Scrollbar(frame_cuerpo, orient=tk.VERTICAL, command=tree.yview)
scrollbar.grid(row=0, column=1, sticky="ns")
tree.configure(yscrollcommand=scrollbar.set)

for col, title in zip(columns, ["Patente", "Sector", "Ingreso", "Egreso", "Tiempo estimado", "Tiempo total", "Precio x Hora", "Precio Total"]):
    tree.heading(col, text=title)
    tree.column(col, anchor=tk.CENTER, width=120)

mostrar_datos()

root.mainloop()
