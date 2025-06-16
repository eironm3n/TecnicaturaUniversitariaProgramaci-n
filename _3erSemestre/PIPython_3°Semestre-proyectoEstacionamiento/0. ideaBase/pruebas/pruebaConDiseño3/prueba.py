import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

# Configuración base de datos
conn = sqlite3.connect('estacionamiento.db')
cursor = conn.cursor()
cursor.execute('''CREATE TABLE IF NOT EXISTS vehiculos (
    patente TEXT,
    sector TEXT,
    ingreso TEXT,
    egreso TEXT,
    tiempo_estimado TEXT,
    tiempo_total TEXT,
    precio_hora REAL,
    precio_total REAL,
    pago_confirmado INTEGER DEFAULT 0
)''')
conn.commit()

# Guardar en Excel
excel_file = "registro_estacionamiento.xlsx"
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["Patente", "Sector", "Ingreso", "Egreso", "Tiempo estimado", "Tiempo total", "Precio por Hora", "Precio Total"])
    wb.save(excel_file)

def actualizar_excel():
    wb = load_workbook(excel_file)
    ws = wb.active
    ws.delete_rows(2, ws.max_row)
    cursor.execute("SELECT * FROM vehiculos")
    for row in cursor.fetchall():
        ws.append(row)
    wb.save(excel_file)

def registrar():
    patente = entry_patente.get()
    sector = combo_sector.get()
    tiempo_est = entry_tiempo.get()
    ingreso = entry_ingreso.get()
    egreso = entry_egreso.get()

    if not patente or not sector or not ingreso:
        messagebox.showerror("Error", "Patente, Sector e Ingreso son obligatorios.")
        return

    cursor.execute("INSERT INTO vehiculos (patente, sector, ingreso, egreso, tiempo_estimado, tiempo_total, precio_hora, precio_total) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                   (patente, sector, ingreso, egreso, tiempo_est, '', 0.0, 0.0))
    conn.commit()
    actualizar_excel()
    mostrar_datos()
    entry_patente.delete(0, tk.END)
    entry_tiempo.delete(0, tk.END)
    entry_ingreso.delete(0, tk.END)
    entry_egreso.delete(0, tk.END)

def mostrar_datos():
    for item in tree.get_children():
        tree.delete(item)
    cursor.execute("SELECT * FROM vehiculos")
    for row in cursor.fetchall():
        color = "turquoise" if row[3] else "white"
        tree.insert("", tk.END, values=row, tags=(color,))
    tree.tag_configure("turquoise", background="turquoise")
    tree.tag_configure("white", background="white")

def actualizar_reloj():
    hora_actual = datetime.now().strftime("%H:%M:%S")
    reloj_label.config(text=hora_actual)
    reloj_label.after(1000, actualizar_reloj)

# GUI
root = tk.Tk()
root.title("Sistema de Estacionamiento")
root.geometry("1200x650")

# Bloque principal superior
frame_superior = tk.Frame(root)
frame_superior.pack(fill=tk.X, padx=10, pady=10)

# Reloj digital moderno en la esquina superior derecha
reloj_label = tk.Label(frame_superior, text="", font=("Segoe UI", 30, "bold"), fg="white", bg="black", padx=20, pady=10)
reloj_label.pack(side=tk.RIGHT, padx=10)
actualizar_reloj()

# Bloques de formulario
frame_bloque1 = tk.Frame(frame_superior)
frame_bloque1.pack(side=tk.LEFT, padx=10)

frame_bloque2 = tk.Frame(frame_superior)
frame_bloque2.pack(side=tk.LEFT, padx=10)

# Etiquetas bloque 1
tk.Label(frame_bloque1, text="Patente", font=("Arial", 10)).grid(row=0, column=0)
tk.Label(frame_bloque1, text="Sector de estacionamiento", font=("Arial", 10)).grid(row=0, column=1)
tk.Label(frame_bloque1, text="Tiempo estimado", font=("Arial", 10)).grid(row=0, column=2)

# Entradas bloque 1
entry_patente = tk.Entry(frame_bloque1, width=15, font=("Arial", 12))
entry_patente.grid(row=1, column=0, padx=5)
entry_patente.insert(0, "")

combo_sector = ttk.Combobox(frame_bloque1, values=["Sector 1", "Sector 2", "Sector 3", "Sector 4"], width=15, font=("Arial", 12))
combo_sector.grid(row=1, column=1, padx=5)
combo_sector.set("")

entry_tiempo = tk.Entry(frame_bloque1, width=15, font=("Arial", 12))
entry_tiempo.grid(row=1, column=2, padx=5)
entry_tiempo.insert(0, "")

# Etiquetas bloque 2
tk.Label(frame_bloque2, text="Ingreso", font=("Arial", 10)).grid(row=0, column=0)
tk.Label(frame_bloque2, text="Egreso", font=("Arial", 10)).grid(row=0, column=1)

# Entradas bloque 2
entry_ingreso = tk.Entry(frame_bloque2, width=15, font=("Arial", 12))
entry_ingreso.grid(row=1, column=0, padx=5)
entry_ingreso.insert(0, "")

entry_egreso = tk.Entry(frame_bloque2, width=15, font=("Arial", 12))
entry_egreso.grid(row=1, column=1, padx=5)
entry_egreso.insert(0, "")

# Botón registrar
btn_registrar = tk.Button(frame_superior, text="Registrar", command=registrar)
btn_registrar.pack(side=tk.LEFT, padx=10)

# Tabla de visualización
columns = ("patente", "sector", "ingreso", "egreso", "tiempo_estimado", "tiempo_total", "precio_hora", "precio_total")
tree = ttk.Treeview(root, columns=columns, show="headings")
tree.pack(fill=tk.BOTH, expand=True)

tree.heading("patente", text="Patente")
tree.heading("sector", text="Sector")
tree.heading("ingreso", text="Ingreso")
tree.heading("egreso", text="Egreso")
tree.heading("tiempo_estimado", text="Tiempo estimado")
tree.heading("tiempo_total", text="Tiempo total")
tree.heading("precio_hora", text="Precio x Hora")
tree.heading("precio_total", text="Precio Total")

mostrar_datos()

root.mainloop()
